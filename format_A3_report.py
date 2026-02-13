import json
from urllib.parse import unquote
from openpyxl import load_workbook, Workbook
import argparse
import os
from datetime import datetime, timedelta
import excel

TEMPLATE_EXCEL_FILE = "NANO_KAFFEE_GmbH_YYYY_MM_Abt.3A.xlsx"


def get_previous_month_range():
    today = datetime.today()
    first_day_of_current_month = today.replace(day=1)
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
    first_day_of_previous_month = last_day_of_previous_month.replace(day=1)

    return first_day_of_previous_month, last_day_of_previous_month


class A3Report:
    def __init__(self, wb: Workbook):
        self.wb = wb
        self.ws = self.wb.active
        self.amount = {"EU": 0, "Ausfuhr": 0}

    def _populate_row(self, raw_data, idx):
        row = excel.row.copy()
        for key, source_key in excel.row_key_to_json_key.items():
            row[key] = int(raw_data[source_key]) / 1000 if key == "Amount" else raw_data[source_key]
        row["ID"] = idx + 1
        row["Region"] = "EU" if "B2B" in raw_data['customer.now.tags'] else "Ausfuhr"
        self.amount[row["Region"]] += row["Amount"]
        self.ws.append(list(row.values()))

    def append_json_to_xlsx(self, json_string):
        self.ws.delete_rows(excel.HEADER_ROW + 1, self.ws.max_row)
        # To properly handle the '-character in property names (e.g. cumstomer_address), the n8n-node has to encodeURI
        # the request body before sending it to the script. Therefore, we have to decode the JSON string before loading.
        decoded_json = unquote(json_string)
        rows = json.loads(decoded_json)
        idx = 0
        for raw_data in rows:
            self._populate_row(raw_data, idx)
            idx += 1

    def save(self, target_path):
        self.ws[excel.AMOUNT["EU"]] = self.amount["EU"]
        self.ws[excel.AMOUNT["Ausfuhr"]] = self.amount["Ausfuhr"]
        start_date, end_date = get_previous_month_range()
        self.ws[excel.TIME_FROM] = start_date.strftime('%d.%m.%Y')
        self.ws[excel.TIME_TO] = end_date.strftime('%d.%m.%Y')
        target_file = TEMPLATE_EXCEL_FILE.replace("YYYY", start_date.strftime('%Y'))
        target_file = target_file.replace("MM", start_date.strftime('%m'))
        # Save the updated XLSX file
        self.wb.save(os.path.join(target_path, target_file))


def main():
    # Create an argument parser
    parser = argparse.ArgumentParser()

    # Add arguments for JSON string and output path
    parser.add_argument("json_string", type=str, help="JSON string input")
    parser.add_argument("excel_path", type=str, help="Path where the Excel files are (template and target)")

    # Parse the arguments
    args = parser.parse_args()

    wb = load_workbook(os.path.join(args.excel_path, TEMPLATE_EXCEL_FILE))
    report = A3Report(wb)
    report.append_json_to_xlsx(args.json_string)
    report.save(args.excel_path)


if __name__ == "__main__":
    main()
