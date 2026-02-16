import xlrd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook, Workbook
import argparse
import os
from datetime import datetime, timedelta
import excel
from excel import A1_TEMPLATE_EXCEL_FILE as TEMPLATE_EXCEL_FILE


def get_previous_month_range():
    today = datetime.today()
    first_day_of_current_month = today.replace(day=1)
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
    first_day_of_previous_month = last_day_of_previous_month.replace(day=1)

    return first_day_of_previous_month, last_day_of_previous_month


def build_excel_file_name(start_date, template_name=TEMPLATE_EXCEL_FILE):
    return template_name.replace("YYYY", start_date.strftime('%Y'))\
        .replace("MM", start_date.strftime('%m'))


class A1Report:
    def __init__(self, wb: Workbook):
        self.wb = wb
        self.ws = self.wb.active
        self._total = 0

    def add_summary(self, a3_sheet: Worksheet):
        self.ws[excel.A1_SUMMARY_TOTAL] = self._total
        self.ws[excel.A1_SUMMARY_EU] = a3_sheet[excel.A3_AMOUNT_EU].value
        self.ws[excel.A1_SUMMARY_AUSFUHR] = a3_sheet[excel.A3_AMOUNT_AUSFUHR].value

    def _parse_row(self, source_row):
        source_row.pop()  # skip weight loss info
        new_row = []
        for cell in source_row:
            if cell.ctype == xlrd.XL_CELL_DATE:
                date_value = xlrd.xldate_as_datetime(cell.value, 0)
                new_row.append(date_value.strftime('%d.%m.%Y %H:%M'))
            else:
                new_row.append(cell.value)
        return new_row

    def add_production_report(self, report_sheet: Worksheet):
        self.ws.delete_rows(excel.A1_HEADER_ROW + 1, self.ws.max_row)
        rows = list(report_sheet.get_rows())
        rows.pop(0)  # skip header row
        for source_row in rows:
            self.ws.append(self._parse_row(source_row))
        # TODO: make computation of totals more robust.
        self._total = float(rows[-1][-1].value)  # Currently fetch totals from the last row of the report

    def save(self, target_path):
        start_date, end_date = get_previous_month_range()
        self.ws[excel.TIME_FROM] = start_date.strftime('%d.%m.%Y')
        self.ws[excel.TIME_TO] = end_date.strftime('%d.%m.%Y')
        target_file = build_excel_file_name(start_date)
        # Save the updated XLSX file
        self.wb.save(os.path.join(target_path, target_file))


def main():
    # Create an argument parser
    parser = argparse.ArgumentParser()

    # Add arguments for JSON string and output path
    parser.add_argument("source_file", type=str, help="Full path to Cropsters' productionReport")
    parser.add_argument("excel_path", type=str, help="Path where the Excel files are (template and target)")

    # Parse the arguments
    args = parser.parse_args()

    wb = load_workbook(os.path.join(args.excel_path, TEMPLATE_EXCEL_FILE))

    production_report = xlrd.open_workbook(args.source_file)
    report = A1Report(wb)
    report.add_production_report(production_report.sheet_by_index(0))

    a3_filename = build_excel_file_name(get_previous_month_range()[0], excel.A3_TEMPLATE_EXCEL_FILE)
    a3_wb = load_workbook(os.path.join(args.excel_path, a3_filename))

    report.add_summary(a3_wb.active)
    report.save(args.excel_path)


if __name__ == "__main__":
    main()
