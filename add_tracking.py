import argparse
import os
from urllib.parse import unquote
import json
from openpyxl import load_workbook, Workbook

import excel
from format_A3_report import build_excel_file_name, get_previous_month_range

TRACKING_URL_PREFIX = "https://www.fedex.com/fedextrack/?trknbr="
ORDER_URL_PREFIX = "https://admin.shopify.com/store/nano-kaffee/orders/"


def create_order_to_tracking_url_mapping(json_string):
    json_string = unquote(json_string)
    rows = json.loads(json_string)

    return {row['name']:
                {"tracking": row['tracking_numbers'][0], "id": row["_id"]}
            for row in rows}


class TrackingManager:
    def __init__(self, wb: Workbook):
        self.wb = wb
        self.ws = self.wb.active

    def add_tracking(self, mapping):
        for row_idx in range(excel.A3_HEADER_ROW + 1, self.ws.max_row + 1):
            order_cell = excel.ORDER_COL + str(row_idx)
            order_id = self.ws[order_cell].value
            if order_id in mapping:
                self.ws[excel.TRACKER_COL + str(row_idx)].value = mapping[order_id]["tracking"]
                self.ws[excel.TRACKER_COL + str(row_idx)].hyperlink = TRACKING_URL_PREFIX + mapping[order_id]["tracking"]
                self.ws[excel.TRACKER_COL + str(row_idx)].style = "Hyperlink"
                self.ws[order_cell].hyperlink = ORDER_URL_PREFIX + mapping[order_id]["id"]
                self.ws[order_cell].style = "Hyperlink"

    def save(self, target_path):
        self.wb.save(target_path)
        print(target_path)


def main():
    # Create an argument parser
    parser = argparse.ArgumentParser()

    # Add arguments for JSON string and output path
    parser.add_argument("json_string", type=str, help="JSON string input")
    parser.add_argument("excel_path", type=str, help="Path where the Excel files are (template and target)")

    # Parse the arguments
    args = parser.parse_args()

    file_name = build_excel_file_name(*get_previous_month_range())
    file_path = os.path.join(args.excel_path, file_name)
    wb = load_workbook(file_path)

    order_2_tracking = create_order_to_tracking_url_mapping(args.json_string)
    tracker = TrackingManager(wb)
    tracker.add_tracking(order_2_tracking)
    tracker.save(file_path)


if __name__ == "__main__":
    main()
